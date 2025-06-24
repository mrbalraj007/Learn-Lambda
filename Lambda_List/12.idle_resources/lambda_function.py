import boto3
import datetime
import io
import os
import json
from contextlib import redirect_stdout
from openpyxl import Workbook

def lambda_handler(event, context):
    buffer = io.StringIO()
    with redirect_stdout(buffer):
        ec2 = boto3.client('ec2')
        iam = boto3.client('iam')
        lambda_client = boto3.client('lambda')
        cloudwatch = boto3.client('cloudwatch')
        cloudtrail = boto3.client('cloudtrail')
        sns = boto3.client('sns')

        now = datetime.datetime.now(datetime.timezone.utc)

        # Helper function to get creator information from CloudTrail
        def get_resource_creator(resource_type, resource_id):
            try:
                # Define event names based on resource type
                event_names = {
                    'ebs': ['CreateVolume'],
                    'snapshot': ['CreateSnapshot', 'CopySnapshot'],
                    'instance': ['RunInstances'],
                    'security-group': ['CreateSecurityGroup'],
                    'role': ['CreateRole'],
                    'lambda': ['CreateFunction']
                }
                
                # Look up events for this resource type
                events = event_names.get(resource_type, [])
                if not events:
                    return "Unknown"
                
                # Get the timeframe for lookup - default to 90 days (CloudTrail limitation)
                lookup_days = 90
                start_time = now - datetime.timedelta(days=lookup_days)
                
                # Lookup in CloudTrail
                response = cloudtrail.lookup_events(
                    LookupAttributes=[
                        {
                            'AttributeKey': 'ResourceName',
                            'AttributeValue': resource_id
                        }
                    ],
                    StartTime=start_time,
                    EndTime=now,
                    MaxResults=10
                )
                
                # Process results
                for event in response.get('Events', []):
                    cloud_trail_event = json.loads(event.get('CloudTrailEvent', '{}'))
                    event_name = cloud_trail_event.get('eventName', '')
                    
                    if event_name in events:
                        username = cloud_trail_event.get('userIdentity', {}).get('userName', '')
                        user_type = cloud_trail_event.get('userIdentity', {}).get('type', '')
                        if not username and user_type:
                            # If userName is not available, use type or principalId
                            username = user_type
                            principalId = cloud_trail_event.get('userIdentity', {}).get('principalId', '')
                            if principalId:
                                username += f" ({principalId})"
                        
                        # Return username or a default if none found
                        return username if username else "System or Service"
                
                return "Unknown"
            except Exception as e:
                print(f"Error retrieving creator for {resource_type} {resource_id}: {str(e)}")
                return "Lookup Error"

        # 1. Idle EBS Volumes
        print("\n[1] Idle EBS Volumes:")
        volumes = ec2.describe_volumes(Filters=[{'Name': 'status', 'Values': ['available']}])['Volumes']
        # Add creator info for volumes
        for volume in volumes:
            volume['CreatedBy'] = get_resource_creator('ebs', volume['VolumeId'])
            
        if volumes:
            for volume in volumes:
                print(f"  - Volume ID: {volume['VolumeId']}, Size: {volume['Size']} GiB, Created By: {volume['CreatedBy']}")
        else:
            print("  - No idle EBS volumes found.")

        # 2. Snapshot Audit
        print("\n[2] Snapshot Expiry Tag Audit:")
        snapshots = ec2.describe_snapshots(OwnerIds=['self'])['Snapshots']
        # Add creator info for snapshots
        for snapshot in snapshots:
            snapshot['CreatedBy'] = get_resource_creator('snapshot', snapshot['SnapshotId'])
            
        flagged_long_expiry = []
        missing_tag = []
        invalid_tag = []

        for snapshot in snapshots:
            tags = snapshot.get('Tags', [])
            expiry_tag = next((tag['Value'] for tag in tags if tag['Key'] == 'ExpiryDate'), None)

            if not expiry_tag:
                missing_tag.append(snapshot)
                continue

            try:
                expiry_date = datetime.datetime.strptime(expiry_tag, "%Y-%m-%d").date()
            except ValueError:
                invalid_tag.append((snapshot, expiry_tag))
                continue

            days_ahead = (expiry_date - datetime.date.today()).days
            if days_ahead > 90:
                flagged_long_expiry.append((snapshot, expiry_date, days_ahead))

        if flagged_long_expiry:
            print("  - Snapshots with ExpiryDate more than 90 days in the future:")
            for snapshot, expiry_date, days in flagged_long_expiry:
                print(f"    • Snapshot ID: {snapshot['SnapshotId']}, Expiry Date: {expiry_date}, {days} days ahead, Size: {snapshot['VolumeSize']} GiB")
        else:
            print("  - No snapshots found with ExpiryDate > 90 days.")

        if invalid_tag:
            print("  - Snapshots with invalid ExpiryDate tag format:")
            for snapshot, tag_value in invalid_tag:
                print(f"    • Snapshot ID: {snapshot['SnapshotId']}, Tag Value: '{tag_value}' (Expected YYYY-MM-DD)")

        if missing_tag:
            print("  - Snapshots missing the 'ExpiryDate' tag:")
            for snapshot in missing_tag:
                print(f"    • Snapshot ID: {snapshot['SnapshotId']}, Size: {snapshot['VolumeSize']} GiB")
        else:
            print("  - All snapshots have the 'ExpiryDate' tag.")

        # 2b. Untagged and Possibly Orphaned Snapshots
        print("\n[2b] Untagged Snapshots Not Linked to Any Volume or EC2 Instance:")
        all_volume_ids = set(vol['VolumeId'] for vol in ec2.describe_volumes()['Volumes'])

        reservations = ec2.describe_instances(Filters=[{'Name': 'instance-state-name', 'Values': ['pending', 'running', 'stopping', 'stopped']}])['Reservations']
        all_instance_ids = set()
        for reservation in reservations:
            for instance in reservation['Instances']:
                all_instance_ids.add(instance['InstanceId'])

        untagged_orphaned_snapshots = []
        for snapshot in snapshots:
            tags = snapshot.get('Tags', [])
            expiry_tag = next((tag['Value'] for tag in tags if tag['Key'] == 'ExpiryDate'), None)

            if expiry_tag:
                continue  # Tagged → skip

            volume_id = snapshot.get('VolumeId')
            if volume_id in all_volume_ids:
                continue  # Volume exists → skip

            # Try to detect EC2 link from description
            description = snapshot.get('Description', '').lower()
            linked_instance = any(instance_id.lower() in description for instance_id in all_instance_ids)

            if not linked_instance:
                untagged_orphaned_snapshots.append(snapshot)

        if untagged_orphaned_snapshots:
            for snap in untagged_orphaned_snapshots:
                print(f"  - Snapshot ID: {snap['SnapshotId']}, Volume ID: {snap.get('VolumeId', 'N/A')}, Created: {snap['StartTime'].date()}, Size: {snap['VolumeSize']} GiB")
        else:
            print("  - No untagged, orphaned snapshots without EC2 link found.")

        # 3. Stopped EC2 Instances
        print("\n[3] Stopped EC2 Instances:")
        instances = ec2.describe_instances(Filters=[{'Name': 'instance-state-name', 'Values': ['stopped']}])['Reservations']
        for reservation in instances:
            for instance in reservation['Instances']:
                instance['CreatedBy'] = get_resource_creator('instance', instance['InstanceId'])
                
        if instances and any(r['Instances'] for r in instances):
            for reservation in instances:
                for instance in reservation['Instances']:
                    print(f"  - Instance ID: {instance['InstanceId']}, Type: {instance['InstanceType']}, Created By: {instance['CreatedBy']}")
        else:
            print("  - No stopped EC2 instances found.")

        # 4. Unused Security Groups
        print("\n[4] Unused Security Groups:")
        all_sgs = ec2.describe_security_groups()['SecurityGroups']
        # Add creator info for security groups
        for sg in all_sgs:
            sg['CreatedBy'] = get_resource_creator('security-group', sg['GroupId'])
            
        used_sgs = set()
        enis = ec2.describe_network_interfaces()['NetworkInterfaces']
        for eni in enis:
            for group in eni['Groups']:
                used_sgs.add(group['GroupId'])

        unused_sgs = [sg for sg in all_sgs if sg['GroupId'] not in used_sgs and sg['GroupName'] != 'default']
        if unused_sgs:
            for sg in unused_sgs:
                print(f"  - Security Group ID: {sg['GroupId']}, Name: {sg['GroupName']}, Created By: {sg['CreatedBy']}")
        else:
            print("  - No unused security groups found.")

        # 5. IAM Roles with No Attached Policies
        print("\n[5] IAM Roles with No Attached Policies:")
        roles = iam.list_roles()['Roles']
        for role in roles:
            role['CreatedBy'] = get_resource_creator('role', role['RoleName'])
            
        found = False
        for role in roles:
            try:
                attached_policies = iam.list_attached_role_policies(RoleName=role['RoleName'])['AttachedPolicies']
                inline_policies = iam.list_role_policies(RoleName=role['RoleName'])['PolicyNames']
            except Exception as e:
                print(f"  - Skipping role {role['RoleName']} due to error: {e}")
                continue

            if not attached_policies and not inline_policies:
                print(f"  - Role Name: {role['RoleName']}, Created By: {role['CreatedBy']}")
                found = True
        if not found:
            print("  - All roles have attached or inline policies.")

        # 6. Idle Lambda Functions (No Invocations in 30+ Days)
        print("\n[6] Idle Lambda Functions (No Invocations in 30+ Days):")
        functions = lambda_client.list_functions()['Functions']
        for function in functions:
            function['CreatedBy'] = get_resource_creator('lambda', function['FunctionName'])
            
        found = False
        for function in functions:
            function_name = function['FunctionName']

            try:
                metrics = cloudwatch.get_metric_statistics(
                    Namespace='AWS/Lambda',
                    MetricName='Invocations',
                    Dimensions=[{'Name': 'FunctionName', 'Value': function_name}],
                    StartTime=now - datetime.timedelta(days=30),
                    EndTime=now,
                    Period=2592000,
                    Statistics=['Sum']
                )
            except Exception as e:
                print(f"  - Could not fetch metrics for {function_name}: {str(e)}")
                continue

            datapoints = metrics.get('Datapoints', [])
            if isinstance(datapoints, list) and datapoints and isinstance(datapoints[0], dict):
                invocation_count = datapoints[0].get('Sum', 0)
            else:
                invocation_count = 0

            if invocation_count == 0:
                print(f"  - Function Name: {function_name}, Created By: {function['CreatedBy']}")
                found = True

        if not found:
            print("  - All functions have been invoked within the last 30 days.")

    # build Excel workbook with separate sheets for each resource
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)
    
    # Idle EBS Volumes sheet
    ws = wb.create_sheet('IdleEBS')
    ws.append(['VolumeId','Status','SizeGiB','Region','CreationDate','CreatedBy'])
    for vol in volumes:
        ws.append([
            vol['VolumeId'],
            vol.get('State','available'),
            vol.get('Size'),
            os.environ.get('REGION',''),
            vol.get('CreateTime').isoformat() if vol.get('CreateTime') else '',
            vol.get('CreatedBy', 'Unknown')
        ])
        
    # Snapshot Expiry sheet
    ws = wb.create_sheet('SnapshotExpiry')
    ws.append(['SnapshotId','ExpiryTag','CreationDate','CreatedBy'])
    for snap in snapshots:
        expiry = next((t['Value'] for t in snap.get('Tags',[]) if t['Key']=='ExpiryDate'), '')
        ws.append([
            snap['SnapshotId'], 
            expiry,
            snap.get('StartTime').isoformat() if snap.get('StartTime') else '',
            snap.get('CreatedBy', 'Unknown')
        ])
        
    # Stopped Instances sheet
    ws = wb.create_sheet('StoppedInstances')
    ws.append(['InstanceId','State','CreationDate','CreatedBy'])
    for r in instances:
        for inst in r['Instances']:
            ws.append([
                inst['InstanceId'], 
                inst['State']['Name'],
                inst.get('LaunchTime').isoformat() if inst.get('LaunchTime') else '',
                inst.get('CreatedBy', 'Unknown')
            ])
            
    # Security Groups sheet
    ws = wb.create_sheet('UnusedSecurityGroups')
    ws.append(['GroupId','GroupName','CreationDate','CreatedBy'])
    for sg in unused_sgs:
        ws.append([
            sg['GroupId'], 
            sg['GroupName'], 
            '',
            sg.get('CreatedBy', 'Unknown')
        ])
        
    # IAM Roles sheet
    ws = wb.create_sheet('IAMRolesNoPolicies')
    ws.append(['RoleName','Description','CreationDate','CreatedBy'])
    for role in roles:
        attached = iam.list_attached_role_policies(RoleName=role['RoleName'])['AttachedPolicies']
        inline = iam.list_role_policies(RoleName=role['RoleName'])['PolicyNames']
        if not attached and not inline:
            ws.append([
                role['RoleName'], 
                role.get('Description',''),
                role.get('CreateDate').isoformat() if role.get('CreateDate') else '',
                role.get('CreatedBy', 'Unknown')
            ])
            
    # Idle Lambdas sheet
    ws = wb.create_sheet('IdleLambdas')
    ws.append(['FunctionName','CreationDate','CreatedBy'])
    for function in functions:
        # Recheck invocation metrics
        metrics = cloudwatch.get_metric_statistics(
            Namespace='AWS/Lambda',
            MetricName='Invocations',
            Dimensions=[{'Name': 'FunctionName', 'Value': function['FunctionName']}],
            StartTime=now - datetime.timedelta(days=30),
            EndTime=now,
            Period=2592000,
            Statistics=['Sum']
        )
        invocation_count = sum(datapoint.get('Sum', 0) for datapoint in metrics.get('Datapoints', []))
        
        if invocation_count == 0:
            ws.append([
                function['FunctionName'],
                function.get('LastModified',''),
                function.get('CreatedBy', 'Unknown')
            ])
            
    # save and upload
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    s3 = boto3.client('s3')
    timestamp = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    key = f"report/{timestamp}.xlsx"
    s3.put_object(
        Bucket=os.environ['OUTPUT_BUCKET'], 
        Key=key, 
        Body=output.getvalue(), 
        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Send SNS notification
    report_url = f"https://s3.console.aws.amazon.com/s3/object/{os.environ['OUTPUT_BUCKET']}/{key}"
    sns_message = f"""
    AWS Resource Audit Report - {timestamp}
    
    Weekly report of potential cost-saving opportunities has been generated.
    
    Report Location: {report_url}
    
    Summary:
    - Idle EBS Volumes: {len(volumes)}
    - Snapshots Missing Expiry Tag: {len([s for s in snapshots if not next((t['Value'] for t in s.get('Tags', []) if t['Key'] == 'ExpiryDate'), None)])}
    - Stopped EC2 Instances: {sum(len(r['Instances']) for r in instances)}
    - Unused Security Groups: {len(unused_sgs)}
    - IAM Roles without Policies: {len([r for r in roles if not iam.list_attached_role_policies(RoleName=r['RoleName'])['AttachedPolicies'] and not iam.list_role_policies(RoleName=r['RoleName'])['PolicyNames']])}
    - Idle Lambda Functions: {len([f for f in functions if not any(d.get('Sum', 0) > 0 for d in cloudwatch.get_metric_statistics(Namespace='AWS/Lambda', MetricName='Invocations', Dimensions=[{'Name': 'FunctionName', 'Value': f['FunctionName']}], StartTime=now - datetime.timedelta(days=30), EndTime=now, Period=2592000, Statistics=['Sum']).get('Datapoints', []))])}
    """
    
    sns.publish(
        TopicArn=os.environ['SNS_TOPIC_ARN'],
        Subject=f"AWS Resource Audit Report - {timestamp}",
        Message=sns_message
    )
    
    print(f"Saved report to s3://{os.environ['OUTPUT_BUCKET']}/{key}")
    print(f"Sent notification to SNS topic: {os.environ['SNS_TOPIC_ARN']}")
    
    return {
        'statusCode': 200,
        'body': json.dumps('Report generated and notification sent')
    }

# Local test entrypoint
if __name__ == '__main__':
    lambda_handler({}, {})
