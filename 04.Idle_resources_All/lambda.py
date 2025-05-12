import datetime
import base64
import os
from io import BytesIO
import boto3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from concurrent.futures import ThreadPoolExecutor, as_completed

def lambda_handler(event, context):
    print("🚀 Lambda execution started")

    # Get S3 bucket from environment variables
    output_bucket = os.environ.get('OUTPUT_BUCKET')
    if not output_bucket:
        raise ValueError("OUTPUT_BUCKET environment variable not set")
    
    # Get AWS region from Lambda's environment
    # This is provided automatically by AWS Lambda
    aws_region = os.environ.get('AWS_REGION')
    print(f"🌐 Running in AWS Region: {aws_region}")

    ec2 = boto3.client('ec2')
    iam = boto3.client('iam')
    lambda_client = boto3.client('lambda')
    s3 = boto3.client('s3')

    timestamp = datetime.datetime.utcnow().strftime('%Y-%m-%d-%H-%M')
    filename = f"idle-resource-report-{timestamp}.xlsx"
    s3_key = f"reports/{filename}"

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    def create_sheet(title, headers):
        safe_title = title.replace("/", "-")
        ws = wb.create_sheet(safe_title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        return ws

    resource_counts = {}

    # Idle EBS Volumes
    print("🔍 Fetching Idle EBS Volumes...")
    volumes = ec2.describe_volumes(Filters=[{'Name': 'status', 'Values': ['available']}])['Volumes']
    print(f"📦 Found {len(volumes)} Idle EBS Volumes")
    ebs_ws = create_sheet("Idle EBS Volumes", ["Volume ID", "Size (GiB)", "Created Time", "Tags"])
    for v in volumes:
        tags = ", ".join(f"{t['Key']}={t['Value']}" for t in v.get('Tags', [])) or "-"
        ebs_ws.append([v['VolumeId'], v['Size'], v['CreateTime'].strftime('%Y-%m-%d %H:%M'), tags])
    resource_counts["Idle EBS Volumes"] = len(volumes)

    # Orphaned Snapshots
    print("🔍 Fetching Orphaned Snapshots...")
    snapshots = ec2.describe_snapshots(OwnerIds=['self'])['Snapshots']
    print(f"📸 Found {len(snapshots)} Snapshots")
    snap_ws = create_sheet("Orphaned Snapshots", ["Snapshot ID", "Volume Size", "ExpiryDate", "Reason"])
    expired_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    expired_count = 0
    for snap in snapshots:
        tags = {t['Key']: t['Value'] for t in snap.get('Tags', [])}
        expiry = tags.get('ExpiryDate')
        is_expired = False
        reason = "No linked volume"
        if expiry:
            try:
                expiry_date = datetime.datetime.strptime(expiry, "%Y-%m-%d")
                if (datetime.datetime.utcnow() - expiry_date).days > 90:
                    is_expired = True
                    reason = "Expired (>90 days)"
            except Exception:
                reason = "Invalid ExpiryDate"
        elif not snap.get("Description", "").startswith("Created by CreateImage"):
            reason = "Orphaned & no expiry tag"

        snap_ws.append([snap['SnapshotId'], snap['VolumeSize'], expiry or "-", reason])
        if is_expired:
            for cell in snap_ws[-1]:
                cell.fill = expired_fill
            expired_count += 1
    resource_counts["Orphaned Snapshots"] = len(snapshots)

    # Stopped EC2 Instances
    print("🔍 Fetching Stopped EC2 Instances...")
    stopped_instances = ec2.describe_instances(Filters=[{'Name': 'instance-state-name', 'Values': ['stopped']}])['Reservations']
    count_stopped = 0
    stopped_ws = create_sheet("Stopped EC2 Instances", ["Instance ID", "Type", "Launch Time", "Tags"])
    for res in stopped_instances:
        for i in res['Instances']:
            tags = ", ".join(f"{t['Key']}={t['Value']}" for t in i.get('Tags', [])) or "-"
            stopped_ws.append([i['InstanceId'], i['InstanceType'], i['LaunchTime'].strftime('%Y-%m-%d %H:%M'), tags])
            count_stopped += 1
    print(f"🛑 Found {count_stopped} Stopped Instances")
    resource_counts["Stopped EC2 Instances"] = count_stopped

    # Unused Security Groups
    print("🔍 Fetching Unused Security Groups...")
    all_sgs = ec2.describe_security_groups()['SecurityGroups']
    used_sgs = {g['GroupId'] for eni in ec2.describe_network_interfaces()['NetworkInterfaces'] for g in eni['Groups']}
    sg_ws = create_sheet("Unused Security Groups", ["Group ID", "Group Name", "Description", "VPC ID"])
    unused_count = 0
    for sg in all_sgs:
        if sg['GroupId'] not in used_sgs and sg['GroupName'] != 'default':
            sg_ws.append([sg['GroupId'], sg['GroupName'], sg.get('Description', ''), sg.get('VpcId', '-')])
            unused_count += 1
    print(f"🔐 Found {unused_count} Unused Security Groups")
    resource_counts["Unused Security Groups"] = unused_count

    # IAM Roles w/o Policies - Optimized with threading
    print("🔍 Fetching IAM Roles without Policies...")
    roles = iam.list_roles()['Roles']
    iam_ws = create_sheet("IAM Roles w-o Policies", ["Role Name", "Created"])

    def role_has_no_policies(role):
        try:
            role_name = role['RoleName']
            attached = iam.list_attached_role_policies(RoleName=role_name)['AttachedPolicies']
            inline = iam.list_role_policies(RoleName=role_name)['PolicyNames']
            if not attached and not inline:
                return role['RoleName'], role['CreateDate']
        except Exception as e:
            print(f"⚠️ Error checking role {role['RoleName']}: {e}")
            return None

    no_policy_roles = 0
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(role_has_no_policies, role) for role in roles]
        for future in as_completed(futures):
            result = future.result()
            if result:
                role_name, create_date = result
                iam_ws.append([role_name, create_date.strftime('%Y-%m-%d')])
                no_policy_roles += 1

    print(f"👤 Found {no_policy_roles} IAM Roles without policies")
    resource_counts["IAM Roles w/o Policies"] = no_policy_roles

    # Idle Lambda Functions
    print("🔍 Fetching Idle Lambda Functions...")
    functions = lambda_client.list_functions()['Functions']
    idle_lambdas = 0
    lambda_ws = create_sheet("Idle Lambda Functions", ["Function Name", "Last Modified", "Runtime"])
    for f in functions:
        last_modified_str = f['LastModified']
        try:
            last_modified = datetime.datetime.strptime(last_modified_str, "%Y-%m-%dT%H:%M:%S.%f%z")
        except ValueError:
            last_modified = datetime.datetime.strptime(last_modified_str, "%Y-%m-%dT%H:%M:%S%z")
        if (datetime.datetime.now(datetime.timezone.utc) - last_modified).days > 30:
            lambda_ws.append([f['FunctionName'], last_modified_str, f['Runtime']])
            idle_lambdas += 1
    print(f"🪂 Found {idle_lambdas} Idle Lambda Functions")
    resource_counts["Idle Lambda Functions"] = idle_lambdas

    # Summary
    print("📝 Generating Summary Sheet...")
    summary_ws.append(["Resource Type", "Count", "Notes"])
    for col in summary_ws[1]:
        col.font = Font(bold=True)
    summary_ws.append(["Idle EBS Volumes", resource_counts["Idle EBS Volumes"], "Available but unattached"])
    summary_ws.append(["Orphaned Snapshots", resource_counts["Orphaned Snapshots"], f"{expired_count} expired ⛔"])
    summary_ws.append(["Stopped EC2 Instances", count_stopped, "Can be reviewed for deletion"])
    summary_ws.append(["Unused Security Groups", unused_count, "No ENI attached"])
    summary_ws.append(["IAM Roles w/o Policies", no_policy_roles, "Unattached to any policy"])
    summary_ws.append(["Idle Lambda Functions", idle_lambdas, "Not used/updated in 30+ days"])

    # Save workbook to memory
    print("💾 Saving Excel file...")
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    # Upload to S3
    print(f"📤 Uploading report to S3 bucket: {output_bucket}, key: {s3_key}")
    s3.upload_fileobj(
        file_stream, 
        output_bucket, 
        s3_key,
        ExtraArgs={
            'ContentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )
    
    # Generate pre-signed URL (valid for 7 days)
    presigned_url = s3.generate_presigned_url(
        'get_object',
        Params={'Bucket': output_bucket, 'Key': s3_key},
        ExpiresIn=604800  # 7 days in seconds
    )
    
    print("✅ Lambda execution completed successfully")
    
    # Return the S3 information and presigned URL
    return {
        'statusCode': 200,
        'headers': {
            'Content-Type': 'application/json'
        },
        'body': {
            'message': 'Report generated successfully',
            's3Location': {
                'bucket': output_bucket,
                'key': s3_key
            },
            'downloadUrl': presigned_url,
            'resourceCounts': resource_counts
        }
    }
